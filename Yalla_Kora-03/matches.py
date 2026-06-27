from playwright.sync_api import sync_playwright
import pandas as pd
from itertools import zip_longest
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,Border,PatternFill,Alignment,Side

def style_sheet(ws):
    back_fill = PatternFill(fill_type="solid",fgColor="282828")
    header_fill = PatternFill(fill_type="solid",fgColor="3C3836")
    white_font_header = Font(name="Courier New",size=14, bold=True, color="EBDBB2")
    white_font = Font(name="Courier New",size=12, bold=True, color="EBDBB2")
    white_border = Border(
        right=Side(color="504945",style="thin"),
        bottom=Side(color="504945",style="thin")
    )
    nment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font_header
        cell.border = white_border
    for i in range(2,ws.max_row + 1):
        for col in ["A","B","C","D","E","F","G"]:
            cell = ws[f"{col}{i}"]
            cell.fill = back_fill
            cell.font = white_font
            cell.border = white_border
            cell.alignment = nment
            #ws[f"B{i}"].alignment = nment
    for column in ws.columns:
        max_length = 0
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2,100)



def to_ex(data,filename):
    df = pd.DataFrame(data)
    df.to_excel(filename,index=False)

team_A = []
team_B = []
channel = []
result_match = []
time = []
status = []
championships = []

date = input("Enter the date (MM/DD/YY): ")
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    page.goto(f"https://www.yallakora.com/matches?date={date}#days")

    #teamA
    page.wait_for_selector("div.title a")
    team_a = page.locator("div.teams.teamA p").all_text_contents()
    team_A.extend(team_a)


    #teamB
    team_b = page.locator("div.teams.teamB p").all_text_contents()
    team_B.extend(team_b)

    #channel
    get_channel = page.locator(".icon-channel").all_text_contents()
    channel.extend(get_channel)

    #result
    result = page.locator(".score").all_text_contents()
    scores = [f"{result[i]} - {result[i+1]}" for i in range(0, len(result), 2)]
    result_match.extend(scores)

    #time
    get_time = page.locator(".time").all_text_contents()
    time.extend(get_time)

    #status
    get_status = page.locator("div.matchStatus span").all_text_contents()
    status.extend(get_status)

    #championship
    tourn_cards = page.locator(".matchCard").all()
    matches_data = []

    for card in tourn_cards:
        championship_name = card.locator(".title h2").inner_text().strip()
        t = card.locator("div.liItem").all_text_contents()
        match_seq = len(t)
        for get_cs in range(match_seq):
            championships.append(championship_name)

    

    browser.close()




all_data = list(zip_longest(championships, team_A, result_match, team_B, status, time, channel, fillvalue="N/A"))

df = pd.DataFrame(all_data,columns=["Championship", "Team 1", "Scores", "Team 2", "Status", "Time", "Channel"])
df.to_excel("matches.xlsx",index=False)
wb = load_workbook("matches.xlsx")
ws = wb.active
ws.title = "Matches"
style_sheet(ws)
wb.save("matches.xlsx")
    


