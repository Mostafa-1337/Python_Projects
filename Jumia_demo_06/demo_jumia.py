from playwright.sync_api import sync_playwright
import pandas as pd
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,Border,PatternFill,Alignment,Side


def style_sheet(ws):
    back_fill = PatternFill(fill_type="solid",fgColor="F4F1EA")
    header_fill = PatternFill(fill_type="solid",fgColor="2C2A29")
    white_font_header = Font(name="Plus Jakarta Sans",size=14, bold=True, color="F4F1EA")
    white_font = Font(name="Plus Jakarta Sans",size=12, bold=True, color="4A4744")
    green_fill = PatternFill(fill_type="solid", start_color="00A30B", end_color="00A30B")
    white_fill = Font(name="Plus Jakarta Sans", size=12, bold=True, color="FFFFFF")
    white_border = Border(
        right=Side(color="E2DDD5",style="thin"),
        bottom=Side(color="E2DDD5",style="thin")
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font_header
        cell.border = white_border
    for i in range(2,ws.max_row + 1):
        for col in ["A","B","C","D","E","F"]:
            cell = ws[f"{col}{i}"]
            cell.fill = back_fill
            cell.font = white_font
            cell.border = white_border
            discount_cell = ws[f"C{i}"]
            try:
                discount_value = int(str(discount_cell.value).replace("%", "").strip())
                if discount_value > 50:
                    discount_cell.fill = green_fill
                    discount_cell.font = white_fill
            except:
                pass
    for column in ws.columns:
        max_length = 0
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 70)



titles = []
prices_after = []
prices_before = []
dollar_price = []
discounts = []
reviews = []
stock_left = []

search = input("Enter What do you need: ")
start_price = input("Start Price: ")
end_price = input("End Price: ")

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    
    page.goto(f"https://www.jumia.com.eg/catalog/?q={search}&price={start_price}-{end_price}#catalog-listing")
    

    links = page.locator("a.core").all()[:10]
    products_link = [link.get_attribute("href") for link in links]
    for l in products_link:

        #tiltes
        page.goto(f"https://www.jumia.com.eg{l}")
        title = page.locator("h1.-fs20.-ptm.-pbxs").inner_text()
        titles.append(title)


        #price after
        price_a = page.locator("span.-b.-ubpt.-tal.-fs24.-prxs").inner_text()
        prices_after.append(price_a)

        #price before
        price_b_el = page.locator("span.-tal.-gy5.-lthr.-fs16.-pvxs.-ubpt").first
        if price_b_el.count() == 0:
            price_b = "Unknown"
            prices_before.append(price_b)
        else:
            price_b = price_b_el.inner_text()
            prices_before.append(price_b)

        #discount
        discount_el = page.locator("span.bdg._dsct._dyn.-mlm").first
        if discount_el.count() == 0:
            discount = "Unknown"
            discounts.append(discount)
        else:
            discount = discount_el.inner_text()
            discounts.append(discount)

        #get_review
        review_el = page.locator("div.stars._m._al").first
        if review_el.count() == 0:
            review = "Unknown"
            reviews.append(review)
        else:
            review = review_el.inner_text()
            reviews.append(review)

        #stock left
        stock_left_el = page.locator("p.-df.-i-ctr.-fs12.-pbm.-rd5").first
        if stock_left_el.count() == 0:
            stock = "In stock"
            stock_left.append(stock)
        else:
            stock = stock_left_el.inner_text()
            stock_left.append(stock)

        
    
    #dollar price
    def to_dollar(x):
        for dollar in x:
            clean_price = dollar.replace("EGP", "")
            clean_price = clean_price.replace(",", "")
            usd_price = float(clean_price) / 50
            dollar_price.append(f"{usd_price:.2f}$")


    browser.close()



#pandas
data = {
    "Product" : titles,
    "Price" : prices_before,
    "Discount" : discounts,
    "Price after discount" : prices_after,
    "Review" : review,
    "Stock left" : stock_left,
}

df = pd.DataFrame(data)
df.to_excel(r"07_jumia\Output.xlsx", index=False)
wb = load_workbook(r"07_jumia\Output.xlsx")
ws = wb.active
ws.title = "Results"
style_sheet(ws)
wb.save(r"07_jumia\Output.xlsx")