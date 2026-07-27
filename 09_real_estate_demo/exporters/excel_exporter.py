import pandas as pd
import openpyxl
from openpyxl.styles import Font,Border,PatternFill,Alignment,Side


def style_sheet(ws):
    back_fill = PatternFill(fill_type="solid",fgColor="F4F1EA")
    header_fill = PatternFill(fill_type="solid",fgColor="2C2A29")
    white_font_header = Font(name="Plus Jakarta Sans",size=14, bold=True, color="F4F1EA")
    white_font = Font(name="Plus Jakarta Sans",size=12, bold=True, color="4A4744")
    green_fill = PatternFill(fill_type="solid", start_color="00A30B", end_color="00A30B")
    white_fill = Font(name="Plus Jakarta Sans", size=12, bold=True, color="FFFFFF")
    white_border = Border(
        right=Side(color="E2DDD5",style="thin"),
        bottom=Side(color="E2DDD5",style="thin")
    )
    nment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    nment_a = Alignment(horizontal="right", vertical="top", wrap_text=True)
    nment_urls = Alignment(horizontal="right", vertical="top", wrap_text=False)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font_header
        cell.border = white_border
    for i in range(2,ws.max_row + 1):
        
        for col in ["A","B","C","D","E","F","G","H","I"]:
            cell = ws[f"{col}{i}"]
            cell.fill = back_fill
            cell.font = white_font
            cell.border = white_border
            cell.alignment = nment

        ws[f"A{i}"].alignment = nment_a
        for al in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            al_text = ws[f"{al}{i}"]
            al_text.alignment = nment

        
        ws[f"I{i}"].alignment = nment_urls
        


    for column in ws.columns:
        max_length = 0
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 40)

def save_to_excel(my_data):
    df = pd.DataFrame(my_data)
    df.to_excel(r"11_real_estate\result.xlsx",index=False)
    wb = openpyxl.load_workbook(r"11_real_estate\result.xlsx")
    ws = wb.active
    ws.title = "Real Estate"
    style_sheet(ws)
    wb.save(r"11_real_estate\result.xlsx")
