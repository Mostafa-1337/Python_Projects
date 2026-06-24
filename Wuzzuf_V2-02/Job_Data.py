#import libraries
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Font,Border,PatternFill,Alignment,Side

##style
def style_sheet(ws):
    back_fill = PatternFill(fill_type="solid",fgColor="282828")
    header_fill = PatternFill(fill_type="solid",fgColor="3C3836")
    white_font_header = Font(name="Courier New",size=14, bold=True, color="EBDBB2")
    white_font = Font(name="Courier New",size=12, bold=True, color="EBDBB2")
    white_border = Border(
        right=Side(color="504945",style="thin"),
        bottom=Side(color="504945",style="thin")
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font_header
        cell.border = white_border
    for i in range(2,ws.max_row + 1):
        for col in ["A","B","C","D","E"]:
            cell = ws[f"{col}{i}"]
            cell.fill = back_fill
            cell.font = white_font
            cell.border = white_border
    for column in ws.columns:
        max_length = 0
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2,50)


#list for lxml
job_title = []
company_name = []
location_name = []
links = []
time = []

#loop to scrap multi pages
for page in range(0,11):
    #request url
    result = requests.get(f"https://wuzzuf.net/search/jobs?q=python&start={page}&a=hpb")
    src = result.content

    #soup
    soup = BeautifulSoup(src,"lxml")
    job_titles = soup.find_all("h2",{"class":"css-193uk2c"})
    company_names = soup.find_all("a",{"class":"css-ipsyv7"})
    location_names = soup.find_all("span",{"class":"css-16x61xq"})
    _time = soup.find_all("a",{"class":"css-a85cz4"})



    #Loops
    for i in range(len(job_titles)):
        job_title.append(job_titles[i].text)
        company_name.append(company_names[i].text)
        location_name.append(location_names[i].text)
        links_path = job_titles[i].find("a").attrs['href']
        links.append(f"https://wuzzuf.net{links_path}")
        time.append(_time[i].text)


    #Write as file with csv
    data = {
        "Job Title":job_title,
        "Comapny Name":company_name,
        "Location":location_name,
        "Link":links,
        "Time":time,
    }
df = pd.DataFrame(data)
df.to_excel("Jobs-V2.xlsx",index=False)

#openpyxel
wb = load_workbook("Jobs-V2.xlsx")
ws = wb.active
ws.title = "Jobs"
style_sheet(ws)
wb.save("Jobs-V2.xlsx")


print("Done!")
print("Check The File")