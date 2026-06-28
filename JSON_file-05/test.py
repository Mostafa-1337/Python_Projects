import requests
import json
import pandas as pd
from openpyxl.styles import Font,Border,PatternFill,Alignment,Side
import openpyxl

def style_sheet(ws):
    back_fill = PatternFill(fill_type="solid",fgColor="282828")
    header_fill = PatternFill(fill_type="solid",fgColor="3C3836")
    white_font_header = Font(name="Courier New",size=14, bold=True, color="EBDBB2")
    white_font = Font(name="Courier New",size=12, bold=True, color="EBDBB2")
    white_border = Border(
        right=Side(color="504945",style="thin"),
        bottom=Side(color="504945",style="thin")
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font_header
        cell.border = white_border
    for i in range(2,ws.max_row + 1):
        for col in ["A","B","C","D","E","F","G"]:
            cell = ws[f"{col}{i}"]
            cell.fill = back_fill
            cell.font = white_font
            cell.border = white_border
    for column in ws.columns:
        max_length = 0
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2,50)

url = "https://dummyjson.com/products"
response = requests.get(url)

def done(text):
    print(f"{text} Done!")

data = response.json()

p_name = []
p_price = []
p_desc = []
p_discount = []
p_ratings = []
p_stock = []
p_brand = []

for item in data['products']:
    p_name.append(item['title'])
    p_price.append(f"{item['price']}$")
    p_desc.append(item['description'])
    p_discount.append(f"%{item['discountPercentage']}")
    p_ratings.append(f"{item['rating']}/5 Stars")
    p_stock.append(item['stock'])
    p_brand.append(item.get('brand','Unknown'))

all_data = {
    "Product":p_name,
    "Price":p_price,
    "Description":p_desc,
    "Discount":p_discount,
    "Ratings":p_ratings,
    "Stock":p_stock,
    "Brand":p_brand
}

#to_excel
df = pd.DataFrame(all_data)
df.to_excel(r"06_json_files\json_data.xlsx",index=False)

wb = openpyxl.load_workbook('json_data.xlsx')
ws = wb.active
ws.title = "Products"
style_sheet(ws)
wb.save(r"06_json_files\json_data.xlsx")