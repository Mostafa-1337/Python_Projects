from playwright.sync_api import sync_playwright
import openpyxl as op
from itertools import zip_longest
from openpyxl.styles import Font,Border,PatternFill,Alignment,Side


back_fill = PatternFill(fill_type="solid",fgColor="282828")
header_fill = PatternFill(fill_type="solid",fgColor="3C3836")
white_font_header = Font(name="Courier New",size=14, bold=True, color="EBDBB2")
white_font = Font(name="Courier New",size=12, bold=True, color="EBDBB2")
white_border = Border(
    right=Side(color="504945",style="thin"),
    bottom=Side(color="504945",style="thin")
)




def add_to_sheet(sheet_name, titles, prices, descs):
    ws = wb.create_sheet(title=sheet_name)
    ws.append(["Product", "Price", "Describtion"])
    for _title, _price, _desc in zip(titles, prices, descs):
        ws.append([_title, _price, _desc])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font_header
        cell.border = white_border
    for i in range(2,ws.max_row + 1):
        for col in ["A","B","C"]:
            cell = ws[f"{col}{i}"]
            cell.fill = back_fill
            cell.font = white_font
            cell.border = white_border
    for column in ws.columns:
        max_length = 0
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2,250)


product_title = []
product_price = []
product_description = []
##########################
product_t = []
product_p = []
product_d = []
##########################
p_title = []
p_price = []
p_description = []


with sync_playwright() as p:
    #go to page
    browser = p.chromium.launch(headless=True)
    page = browser.new_page()
    page.goto("https://webscraper.io/test-sites/e-commerce/allinone")
    page.wait_for_load_state()

    #get product link



    def get_products(url_list, list1, list2, list3):

        for product in url_list:

            #get title
            page.goto(f"https://webscraper.io{product}")
            title = page.locator(".title.card-title").inner_text()
            list1.append(title)


            #get price
            price = page.locator("h4.price.float-end.pull-right span").inner_text()
            list2.append(price)
            

            #get des
            desc = page.locator(".description.card-text").inner_text()
            list3.append(desc)


    ####laptops
    page.click("span:has-text('computers')")
    page.click("span:has-text('laptops')")
    links = page.locator("a.title").all()
    url = [link.get_attribute("href") for link in links]

    get_products(url, product_title, product_price, product_description)


    ###tablets
    page.click("span:has-text('computers')")
    page.click("span:has-text('tablets')")
    links2 = page.locator("a.title").all()
    url2 = [link2.get_attribute("href") for link2 in links2]

    get_products(url2, product_t, product_p, product_d)

    page.click("span:has-text('phones')")
    page.click("span:has-text('touch')")
    links3 = page.locator("a.title").all()
    url3 = [link3.get_attribute("href") for link3 in links3]

    get_products(url3, p_title, p_price, p_description)
    
    

    browser.close()

wb = op.Workbook()
wb.remove(wb.active)

add_to_sheet("Laptops", product_title, product_price, product_description)
add_to_sheet("Tablets", product_t, product_p, product_d)
add_to_sheet("Phones", p_title, p_price, p_description)


wb.save("Big_Data.xlsx")
